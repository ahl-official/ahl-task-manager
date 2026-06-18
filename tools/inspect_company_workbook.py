from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def non_empty_count(row):
    return sum(1 for value in row if value not in (None, ""))


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage: inspect_company_workbook.py <workbook.xlsx>")

    source = Path(sys.argv[1])
    wb = load_workbook(source, read_only=True, data_only=True)
    summary = []

    for ws in wb.worksheets:
        header_row = None
        header_values = []
        sample_rows = []

        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = ["" if value is None else str(value).strip() for value in row]
            if header_row is None and non_empty_count(values) >= 2:
                header_row = idx
                header_values = values
                continue
            if header_row is not None and non_empty_count(values) > 0:
                sample_rows.append(values)
            if len(sample_rows) >= 5:
                break

        summary.append(
            {
                "sheet": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "header_row": header_row,
                "headers": header_values,
                "sample_rows": sample_rows,
            }
        )

    print(json.dumps(summary, indent=2, default=str))


if __name__ == "__main__":
    main()
