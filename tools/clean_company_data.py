from __future__ import annotations

import hashlib
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


SOURCE_TASK_SHEET = "Master"
SOURCE_USER_SHEET = "Doer List"
OUTPUT_DIR = Path("outputs/data-cleaning")


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_text(value: Any) -> str:
    text = as_text(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def normalize_name(value: Any) -> str:
    text = as_text(value).lower()
    text = re.sub(r"[_\-]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    aliases = {
        "vinitt sir": "vinit sir",
        "vinit sir": "vinit sir",
        "vinit": "vinit sir",
        "daizy": "daizy",
    }
    return aliases.get(text, text)


def display_name(value: str) -> str:
    if not value:
        return ""
    if value.endswith(" sir"):
        return value[:-4].title() + " Sir"
    return value.title()


def slug(value: str) -> str:
    text = normalize_name(value)
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "unknown"


def normalize_phone(value: Any) -> str:
    text = as_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    if len(digits) == 10:
        return "91" + digits
    return digits


def normalize_list_value(value: Any) -> str:
    return re.sub(r"\s+", " ", as_text(value)).strip()


def normalize_list_key(value: Any) -> str:
    return normalize_list_value(value).lower()


def load_simple_list(path: Path | None) -> list[str]:
    if not path:
        return []
    values = []
    for line in path.read_text(encoding="utf-8").splitlines():
        text = line.strip()
        if not text or text.startswith("#"):
            continue
        values.append(normalize_list_value(text))
    return values


def parse_args(argv: list[str]) -> tuple[Path, Path | None, Path | None, Path | None]:
    if len(argv) < 2:
        raise SystemExit(
            "Usage: clean_company_data.py <workbook.xlsx> [--departments=departments.txt] [--active-users=users.txt] [--user-departments=user_departments.txt]"
        )

    source = Path(argv[1])
    departments_path = None
    active_users_path = None
    user_departments_path = None
    for arg in argv[2:]:
        if arg.startswith("--departments="):
            departments_path = Path(arg.split("=", 1)[1])
        elif arg.startswith("--active-users="):
            active_users_path = Path(arg.split("=", 1)[1])
        elif arg.startswith("--user-departments="):
            user_departments_path = Path(arg.split("=", 1)[1])
        else:
            raise SystemExit(f"Unknown argument: {arg}")
    return source, departments_path, active_users_path, user_departments_path


def normalize_department(value: Any, allowed_departments: list[str]) -> str:
    raw = normalize_list_value(value)
    aliases = {
        "tecnhician": "Technician",
        "technician": "Technician",
        "editor": "Editor",
        "content creation": "Content Creation",
        "content creation ": "Content Creation",
        "graphic designing": "Graphic Designer",
        "graphic designer": "Graphic Designer",
        "process coordinator": "Process Coordinator",
        "performance marketing": "Performance Marketing",
        "desk incharge": "Desk Incharge",
        "floor manager": "Floor Manager",
        "salon manager": "Salon Manager",
        "colour inventory": "Colour Inventory",
        "customer support": "Customer Support",
        "hair stylist": "Hair Stylist",
        "microbalding": "Microbalding",
        "alchemane manager": "Alchemane Manager",
        "ui/ux": "UI/UX",
        "ai": "AI",
        "hr": "HR",
        "mis": "MIS",
        "crm": "CRM",
    }
    canonical = aliases.get(raw.lower(), raw)
    allowed_by_key = {normalize_list_key(department): department for department in allowed_departments}
    return allowed_by_key.get(normalize_list_key(canonical), canonical)


def parse_user_department_line(line: str, allowed_departments: list[str]) -> tuple[str, str, str] | None:
    text = line.strip()
    if not text or text.startswith("#"):
        return None
    if "-" not in text:
        return None
    name, department = text.split("-", 1)
    phone_match = re.search(r"\(([^)]+)\)", department)
    phone = normalize_phone(phone_match.group(1)) if phone_match else ""
    department = re.sub(r"\([^)]*\)", "", department).strip()
    return normalize_name(name), normalize_department(department, allowed_departments), phone


def load_user_departments(path: Path | None, allowed_departments: list[str]) -> dict[str, dict[str, str]]:
    if not path:
        return {}
    rows: dict[str, dict[str, str]] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        parsed = parse_user_department_line(line, allowed_departments)
        if not parsed:
            continue
        name, department, phone = parsed
        rows[name] = {"department": department, "phone": phone}
    return rows


def phone_last10(value: str) -> str:
    return normalize_phone(value)[-10:]


def normalize_status(value: Any) -> str:
    text = as_text(value).lower()
    if text in {"completed", "complete", "done", "true"}:
        return "Completed"
    if text == "verified":
        return "Verified"
    if text == "overdue":
        return "Overdue"
    if text in {"shifted", "delay requested", "delayed"}:
        return "Delay Requested"
    if text in {"in progress", "in-progress", "started"}:
        return "In Progress"
    return "Pending Accept"


def normalize_priority(value: Any) -> str:
    text = as_text(value).lower()
    if "high" in text or "red" in text or "🔴" in text:
        return "High"
    if "low" in text or "green" in text or "blue" in text or "🔵" in text:
        return "Low"
    return "Medium"


def to_iso_date(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = as_text(value)
    if not text or text.startswith("#"):
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text.split(".")[0], fmt).date().isoformat()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text.split(".")[0]).date().isoformat()
    except ValueError:
        return ""


def stable_task_id(name: str, task: str, first_date: str, final_date: str) -> str:
    seed = f"{normalize_name(name)}|{task.strip().lower()}|{first_date}|{final_date}"
    digest = hashlib.sha1(seed.encode("utf-8")).hexdigest()[:10].upper()
    return f"IMP-{digest}"


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        values = [as_text(cell.value) for cell in col[: min(len(col), 200)]]
        width = max([len(v) for v in values] + [10])
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, 12), 55)


def write_rows(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    style_sheet(ws)


def main() -> None:
    source, departments_path, active_users_path, user_departments_path = parse_args(sys.argv)
    allowed_departments = load_simple_list(departments_path)
    allowed_department_keys = {normalize_list_key(department) for department in allowed_departments}
    user_departments = load_user_departments(user_departments_path, allowed_departments)
    active_user_values = load_simple_list(active_users_path)
    active_user_keys = {normalize_name(value) for value in active_user_values} | set(user_departments)
    active_user_last10 = {phone_last10(value) for value in active_user_values if phone_last10(value)}
    active_user_last10 |= {phone_last10(row["phone"]) for row in user_departments.values() if phone_last10(row["phone"])}
    wb = load_workbook(source, data_only=True, read_only=True)
    if SOURCE_TASK_SHEET not in wb.sheetnames:
        raise SystemExit(f"Missing sheet: {SOURCE_TASK_SHEET}")
    if SOURCE_USER_SHEET not in wb.sheetnames:
        raise SystemExit(f"Missing sheet: {SOURCE_USER_SHEET}")

    users_by_name: dict[str, dict[str, Any]] = {}
    user_rows = list(wb[SOURCE_USER_SHEET].iter_rows(values_only=True))
    for idx, row in enumerate(user_rows[1:], start=2):
        email, name, phone = (list(row) + ["", "", ""])[:3]
        canonical = normalize_name(name)
        if not canonical:
            continue
        phone_clean = normalize_phone(phone)
        users_by_name[canonical] = {
            "uid": f"user-{slug(canonical)}",
            "name": display_name(canonical),
            "raw_name": as_text(name),
            "email": as_text(email).lower(),
            "wa_number": phone_clean,
            "wa_last10": phone_last10(phone_clean),
            "role": "member",
            "department": "",
            "is_active": True,
            "source_sheet": SOURCE_USER_SHEET,
            "source_row": idx,
        }

    for canonical, mapping in user_departments.items():
        if canonical in users_by_name:
            if mapping.get("phone") and not users_by_name[canonical].get("wa_number"):
                users_by_name[canonical]["wa_number"] = mapping["phone"]
                users_by_name[canonical]["wa_last10"] = phone_last10(mapping["phone"])
            users_by_name[canonical]["department"] = mapping["department"]
            continue

        phone = mapping.get("phone", "")
        users_by_name[canonical] = {
            "uid": f"user-{slug(canonical)}",
            "name": display_name(canonical),
            "raw_name": canonical,
            "email": "",
            "wa_number": phone,
            "wa_last10": phone_last10(phone),
            "role": "member",
            "department": mapping["department"],
            "is_active": True,
            "source_sheet": "User Department Mapping",
            "source_row": "",
        }

    task_rows = []
    removed_task_rows = []
    issues = []
    duplicate_counter = Counter()
    status_counter = Counter()
    priority_counter = Counter()
    assignee_counter = Counter()

    ws = wb[SOURCE_TASK_SHEET]
    for source_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cells = list(row) + [""] * 30
        raw_name = cells[0]
        raw_task = cells[1]
        if not as_text(raw_name) and not as_text(raw_task):
            continue

        canonical_name = normalize_name(raw_name)
        task = clean_text(raw_task)
        first_date = to_iso_date(cells[2])
        revision_1 = to_iso_date(cells[3])
        revision_2 = to_iso_date(cells[4])
        final_date = to_iso_date(cells[5])
        status = normalize_status(cells[7])
        remarks = clean_text(cells[11])
        priority = normalize_priority(cells[12])
        notes = clean_text(cells[13])
        dependent_on = clean_text(cells[15])
        dependent_task_id = as_text(cells[16])
        checked_by = clean_text(cells[18])
        assigned_date = to_iso_date(cells[19]) or first_date
        actual_start_date = to_iso_date(cells[20]) or first_date

        user = users_by_name.get(canonical_name)
        uid = user["uid"] if user else f"import-user-{slug(canonical_name)}"
        display = user["name"] if user else display_name(canonical_name)
        wa_number = user["wa_number"] if user else ""
        task_id = stable_task_id(canonical_name, task, first_date, final_date)
        duplicate_key = (canonical_name, task.lower(), first_date, final_date)
        duplicate_counter[duplicate_key] += 1

        row_issues = []
        if not canonical_name:
            row_issues.append("Missing assignee name")
        if not task:
            row_issues.append("Missing task description")
        if canonical_name and not user:
            row_issues.append("Assignee not found in Doer List")
        if user and not wa_number:
            row_issues.append("Assignee missing WhatsApp number")
        if not first_date:
            row_issues.append("Missing/invalid first date")
        if not final_date:
            row_issues.append("Missing/invalid final date")
        if final_date and first_date and final_date < first_date:
            row_issues.append("Final date before first date")

        cleaned = {
            "task_id": task_id,
            "source_sheet": SOURCE_TASK_SHEET,
            "source_row": source_row,
            "assigned_to_uid": uid,
            "assigned_to_name": display,
            "assigned_to_raw_name": as_text(raw_name),
            "assigned_to_wa": wa_number,
            "description": task,
            "category": "One Time",
            "priority": priority,
            "status": status,
            "first_date": first_date,
            "revision_1": revision_1,
            "revision_2": revision_2,
            "final_date": final_date,
            "assigned_date": assigned_date,
            "actual_start_date": actual_start_date,
            "completed_at": final_date if status in {"Completed", "Verified"} else "",
            "remarks": remarks,
            "notes": notes,
            "dependent_on": dependent_on,
            "dependent_task_id": dependent_task_id,
            "checked_by_auditor": checked_by,
            "issue_count": len(row_issues),
            "issues": "; ".join(row_issues),
        }
        task_rows.append(cleaned)
        status_counter[status] += 1
        priority_counter[priority] += 1
        assignee_counter[display or as_text(raw_name) or "Unknown"] += 1
        for issue in row_issues:
            issues.append(
                {
                    "severity": "High" if issue.startswith("Missing task") or issue.startswith("Missing assignee") else "Medium",
                    "sheet": SOURCE_TASK_SHEET,
                    "row": source_row,
                    "field": issue,
                    "assignee": display or as_text(raw_name),
                    "task_preview": task[:120],
                }
            )

    duplicate_issues = []
    seen_duplicate_rows = defaultdict(list)
    for row in task_rows:
        key = (
            normalize_name(row["assigned_to_raw_name"]),
            row["description"].lower(),
            row["first_date"],
            row["final_date"],
        )
        if duplicate_counter[key] > 1:
            seen_duplicate_rows[key].append(row["source_row"])
    for key, rows in seen_duplicate_rows.items():
        duplicate_issues.append(
            {
                "severity": "Medium",
                "sheet": SOURCE_TASK_SHEET,
                "row": ", ".join(map(str, rows)),
                "field": "Possible duplicate task",
                "assignee": display_name(key[0]),
                "task_preview": key[1][:120],
            }
        )

    users_from_tasks = set(normalize_name(row["assigned_to_raw_name"]) for row in task_rows if row["assigned_to_raw_name"])
    for canonical in sorted(users_from_tasks - set(users_by_name)):
        users_by_name[canonical] = {
            "uid": f"import-user-{slug(canonical)}",
            "name": display_name(canonical),
            "raw_name": canonical,
            "email": "",
            "wa_number": "",
            "wa_last10": "",
            "role": "member",
            "department": "",
            "is_active": True,
            "source_sheet": SOURCE_TASK_SHEET,
            "source_row": "",
        }

    users_with_numbers = {
        canonical
        for canonical, user in users_by_name.items()
        if normalize_phone(user.get("wa_number"))
    }
    if active_user_keys or active_user_last10:
        users_with_numbers = {
            canonical
            for canonical in users_with_numbers
            if canonical in active_user_keys or phone_last10(users_by_name[canonical].get("wa_number", "")) in active_user_last10
        }

    for user in users_by_name.values():
        department = normalize_list_value(user.get("department", ""))
        if allowed_department_keys and normalize_list_key(department) not in allowed_department_keys:
            user["department"] = ""
    deleted_users = [
        {
            **user,
            "delete_reason": (
                "User not in active user list"
                if (active_user_keys or active_user_last10) and normalize_phone(user.get("wa_number")) else
                "Missing WhatsApp number"
            ),
        }
        for canonical, user in sorted(users_by_name.items())
        if canonical not in users_with_numbers
    ]
    kept_users = [
        user
        for canonical, user in sorted(users_by_name.items())
        if canonical in users_with_numbers
    ]

    strict_task_rows = []
    for row in task_rows:
        assignee_key = normalize_name(row["assigned_to_raw_name"])
        if assignee_key not in users_with_numbers:
            removed_task_rows.append({
                **row,
                "delete_reason": (
                    "Assigned user no longer exists"
                    if (active_user_keys or active_user_last10) and normalize_phone(row.get("assigned_to_wa")) else
                    "Assigned user has no WhatsApp number"
                ),
            })
        else:
            strict_task_rows.append(row)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out = Workbook()
    out.remove(out.active)

    summary_rows = [
        {"metric": "Source workbook", "value": source.name},
        {"metric": "Task source sheet", "value": SOURCE_TASK_SHEET},
        {"metric": "Original cleaned task rows", "value": len(task_rows)},
        {"metric": "Strict kept task rows", "value": len(strict_task_rows)},
        {"metric": "Removed task rows", "value": len(removed_task_rows)},
        {"metric": "Original cleaned users", "value": len(users_by_name)},
        {"metric": "Strict kept users", "value": len(kept_users)},
        {"metric": "Deleted users without WhatsApp", "value": len(deleted_users)},
        {"metric": "Allowed departments", "value": ", ".join(allowed_departments) if allowed_departments else "No department filter supplied"},
        {"metric": "Active user filter", "value": f"{len(active_user_values)} entries" if active_user_values else "Doer List only"},
        {"metric": "Rows with issues", "value": sum(1 for row in task_rows if row["issue_count"])},
        {"metric": "Possible duplicate groups", "value": len(duplicate_issues)},
        {"metric": "Completed tasks kept", "value": sum(1 for row in strict_task_rows if row["status"] == "Completed")},
        {"metric": "Pending Accept tasks kept", "value": sum(1 for row in strict_task_rows if row["status"] == "Pending Accept")},
        {"metric": "Delay Requested tasks kept", "value": sum(1 for row in strict_task_rows if row["status"] == "Delay Requested")},
    ]

    write_rows(out.create_sheet("Summary"), ["metric", "value"], summary_rows)
    write_rows(
        out.create_sheet("Cleaned Tasks"),
        [
            "task_id",
            "source_sheet",
            "source_row",
            "assigned_to_uid",
            "assigned_to_name",
            "assigned_to_raw_name",
            "assigned_to_wa",
            "description",
            "category",
            "priority",
            "status",
            "first_date",
            "revision_1",
            "revision_2",
            "final_date",
            "assigned_date",
            "actual_start_date",
            "completed_at",
            "remarks",
            "notes",
            "dependent_on",
            "dependent_task_id",
            "checked_by_auditor",
            "issue_count",
            "issues",
        ],
        strict_task_rows,
    )
    write_rows(
        out.create_sheet("Cleaned Users"),
        [
            "uid",
            "name",
            "raw_name",
            "email",
            "wa_number",
            "wa_last10",
            "role",
            "department",
            "is_active",
            "source_sheet",
            "source_row",
        ],
        kept_users,
    )
    write_rows(
        out.create_sheet("Deleted Users No Number"),
        [
            "uid",
            "name",
            "raw_name",
            "email",
            "wa_number",
            "wa_last10",
            "role",
            "department",
            "is_active",
            "source_sheet",
            "source_row",
            "delete_reason",
        ],
        deleted_users,
    )
    write_rows(
        out.create_sheet("Deleted Tasks"),
        [
            "task_id",
            "source_sheet",
            "source_row",
            "assigned_to_uid",
            "assigned_to_name",
            "assigned_to_raw_name",
            "assigned_to_wa",
            "description",
            "category",
            "priority",
            "status",
            "first_date",
            "revision_1",
            "revision_2",
            "final_date",
            "assigned_date",
            "actual_start_date",
            "completed_at",
            "remarks",
            "notes",
            "dependent_on",
            "dependent_task_id",
            "checked_by_auditor",
            "issue_count",
            "issues",
            "delete_reason",
        ],
        removed_task_rows,
    )
    write_rows(
        out.create_sheet("Issues"),
        ["severity", "sheet", "row", "field", "assignee", "task_preview"],
        issues + duplicate_issues,
    )
    write_rows(
        out.create_sheet("Status Summary"),
        ["status", "count"],
        [{"status": key, "count": value} for key, value in status_counter.most_common()],
    )
    write_rows(
        out.create_sheet("Priority Summary"),
        ["priority", "count"],
        [{"priority": key, "count": value} for key, value in priority_counter.most_common()],
    )
    write_rows(
        out.create_sheet("Assignee Summary"),
        ["assignee", "task_count"],
        [{"assignee": key, "task_count": value} for key, value in assignee_counter.most_common()],
    )
    write_rows(
        out.create_sheet("Allowed Departments"),
        ["department"],
        [{"department": department} for department in allowed_departments],
    )

    output_path = OUTPUT_DIR / "ahl_company_data_strict_cleaned_preview.xlsx"
    out.save(output_path)
    print(output_path.resolve())


if __name__ == "__main__":
    main()
