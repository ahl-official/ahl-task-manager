from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def clean_text(value) -> str:
    return str(value or "").strip()


def clean_bool(value) -> bool:
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def read_sheet(ws):
    rows = ws.iter_rows(values_only=True)
    headers = [clean_text(cell) for cell in next(rows)]
    for row in rows:
        record = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
        if any(value not in (None, "") for value in record.values()):
            yield record


def load_departments(path: Path | None) -> list[str]:
    if not path:
        return []
    rows = []
    for line in path.read_text(encoding="utf-8").splitlines():
        text = line.strip()
        if text and not text.startswith("#"):
            rows.append(" ".join(text.split()))
    return rows


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit(
            "Usage: export-cleaned-firestore-json.py <cleaned-workbook.xlsx> [output.json] [--departments=departments.txt]"
        )

    workbook_path = Path(sys.argv[1])
    output_path = Path("outputs/data-cleaning/firestore-import.json")
    departments_path = None
    for arg in sys.argv[2:]:
        if arg.startswith("--departments="):
            departments_path = Path(arg.split("=", 1)[1])
        else:
            output_path = Path(arg)

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    users_ws = wb["Cleaned Users"]
    tasks_ws = wb["Cleaned Tasks"]

    users = []
    for row in read_sheet(users_ws):
        users.append({
            "uid": clean_text(row.get("uid")),
            "name": clean_text(row.get("name")),
            "rawName": clean_text(row.get("raw_name")),
            "email": clean_text(row.get("email")),
            "waNumber": clean_text(row.get("wa_number")),
            "waNumberLast10": clean_text(row.get("wa_last10")),
            "role": clean_text(row.get("role")) or "member",
            "department": clean_text(row.get("department")),
            "isActive": clean_bool(row.get("is_active")),
        })

    tasks = []
    for row in read_sheet(tasks_ws):
        tasks.append({
            "taskId": clean_text(row.get("task_id")),
            "assignedTo": clean_text(row.get("assigned_to_uid")),
            "assignedToName": clean_text(row.get("assigned_to_name")),
            "assignedToWa": clean_text(row.get("assigned_to_wa")),
            "description": clean_text(row.get("description")),
            "category": clean_text(row.get("category")) or "One Time",
            "priority": clean_text(row.get("priority")) or "Medium",
            "status": clean_text(row.get("status")) or "Pending Accept",
            "firstDate": clean_text(row.get("first_date")),
            "revision1": clean_text(row.get("revision_1")),
            "revision2": clean_text(row.get("revision_2")),
            "finalDate": clean_text(row.get("final_date")),
            "assignedDate": clean_text(row.get("assigned_date")),
            "actualStartDate": clean_text(row.get("actual_start_date")),
            "completedAt": clean_text(row.get("completed_at")),
            "remarks": clean_text(row.get("remarks")),
            "notes": clean_text(row.get("notes")),
        })

    payload = {
        "sourceWorkbook": str(workbook_path),
        "allowedDepartments": load_departments(departments_path),
        "users": users,
        "tasks": tasks,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {output_path}")
    print(f"Users: {len(users)}")
    print(f"Tasks: {len(tasks)}")


if __name__ == "__main__":
    main()
